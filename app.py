import streamlit as st
import pandas as pd
import numpy as np
import re
import io
import base64
import warnings
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyBboxPatch
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

# ── MARIANO EVENTS BRAND COLORS ───────────────────────────────
GOLD   = '#F0B43C'
PURPLE = '#8C1478'
TEAL   = '#288C8C'
BLACK  = '#0A0A0A'
DARK   = '#141414'
CARD   = '#1C1C1C'
BORDER = '#2A2A2A'
WHITE  = '#FFFFFF'
LGRAY  = '#AAAAAA'
GREEN  = '#2ECC71'
RED    = '#E74C3C'
PALE_G = '#1A2E1A'
PALE_R = '#2E1A1A'

PALETTE = [GOLD, TEAL, PURPLE, '#F0D080', '#50B0B0', '#B050A0',
           '#F8E4A0', '#80D0D0', '#D080C0', GREEN, RED, '#3498DB']

# ── PAGE CONFIG ───────────────────────────────────────────────
st.set_page_config(
    page_title="Mariano Events — Campaign Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── LOAD LOGO ─────────────────────────────────────────────────
import os
def get_logo_b64():
    logo_path = os.path.join(os.path.dirname(__file__), 'me_logo.png')
    if os.path.exists(logo_path):
        with open(logo_path,'rb') as f:
            return base64.b64encode(f.read()).decode()
    return None
logo_b64 = get_logo_b64()

# ── CSS ───────────────────────────────────────────────────────
st.markdown(f"""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700;900&display=swap');

  html, body, [class*="css"] {{
    background-color: {BLACK};
    color: {WHITE};
    font-family: 'Inter', sans-serif;
  }}
  .main {{ background-color: {BLACK}; }}
  .block-container {{ padding: 0.8rem 1.8rem 2rem 1.8rem; }}

  /* Sidebar */
  section[data-testid="stSidebar"] {{
    background: linear-gradient(180deg, #0F0F0F 0%, #141414 100%);
    border-right: 1px solid {BORDER};
  }}
  section[data-testid="stSidebar"] * {{ color: {WHITE} !important; }}
  section[data-testid="stSidebar"] .stSelectbox > div,
  section[data-testid="stSidebar"] .stMultiSelect > div {{
    background: {DARK} !important;
    border: 1px solid {BORDER} !important;
    border-radius: 8px !important;
  }}

  /* Header bar */
  .me-header {{
    background: linear-gradient(135deg, {DARK} 0%, #1A1A1A 100%);
    border-bottom: 2px solid {GOLD};
    padding: 14px 24px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin: -0.8rem -1.8rem 1.2rem -1.8rem;
    border-radius: 0;
  }}
  .me-brand {{ display: flex; align-items: center; gap: 14px; }}
  .me-title {{
    font-size: 1.4rem; font-weight: 900; letter-spacing: 2px;
    background: linear-gradient(135deg, {GOLD}, {TEAL});
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
  }}
  .me-subtitle {{ font-size: 0.72rem; color: {LGRAY}; letter-spacing: 1px; margin-top: 2px; }}

  /* KPI Cards */
  .kpi-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin-bottom: 16px; }}
  .kpi-card {{
    background: {CARD};
    border: 1px solid {BORDER};
    border-radius: 12px;
    padding: 16px 14px;
    position: relative;
    overflow: hidden;
  }}
  .kpi-card::before {{
    content: '';
    position: absolute; top: 0; left: 0; right: 0; height: 3px;
    background: linear-gradient(90deg, {GOLD}, {TEAL});
  }}
  .kpi-card.purple::before {{ background: linear-gradient(90deg, {PURPLE}, {TEAL}); }}
  .kpi-card.teal::before {{ background: linear-gradient(90deg, {TEAL}, {GOLD}); }}
  .kpi-card.green::before {{ background: linear-gradient(90deg, {GREEN}, {TEAL}); }}
  .kpi-card.red::before {{ background: {RED}; }}
  .kpi-val {{ font-size: 1.75rem; font-weight: 900; color: {GOLD}; line-height: 1.1; }}
  .kpi-val.purple {{ color: #C060A0; }}
  .kpi-val.teal {{ color: {TEAL}; }}
  .kpi-val.green {{ color: {GREEN}; }}
  .kpi-val.red {{ color: {RED}; }}
  .kpi-lbl {{ font-size: 0.65rem; color: {LGRAY}; letter-spacing: 1.5px; text-transform: uppercase; margin-top: 4px; font-weight: 600; }}
  .kpi-sub {{ font-size: 0.7rem; color: #555; margin-top: 2px; }}

  /* Section headers */
  .sec-head {{
    display: flex; align-items: center; gap: 10px;
    margin: 20px 0 10px 0; padding: 10px 14px;
    background: {DARK};
    border-left: 3px solid {GOLD};
    border-radius: 0 8px 8px 0;
  }}
  .sec-head.purple {{ border-left-color: {PURPLE}; }}
  .sec-head.teal {{ border-left-color: {TEAL}; }}
  .sec-head.green {{ border-left-color: {GREEN}; }}
  .sec-head.red {{ border-left-color: {RED}; }}
  .sec-title {{ font-size: 0.95rem; font-weight: 700; color: {GOLD}; letter-spacing: 1px; }}
  .sec-title.purple {{ color: #C060A0; }}
  .sec-title.teal {{ color: {TEAL}; }}
  .sec-title.green {{ color: {GREEN}; }}
  .sec-title.red {{ color: {RED}; }}

  /* Stock alert cards */
  .stock-card {{
    background: {PALE_R};
    border: 1px solid {RED}44;
    border-left: 3px solid {RED};
    border-radius: 8px; padding: 10px 14px; margin-bottom: 8px;
  }}
  .stock-card.warn {{
    background: #2E2A1A;
    border: 1px solid {GOLD}44;
    border-left: 3px solid {GOLD};
  }}
  .stock-card.ok {{
    background: {PALE_G};
    border: 1px solid {GREEN}44;
    border-left: 3px solid {GREEN};
  }}
  .stock-title {{ font-size: 0.9rem; font-weight: 700; color: {WHITE}; }}
  .stock-detail {{ font-size: 0.78rem; color: {LGRAY}; margin-top: 3px; }}

  /* Tabs */
  .stTabs [data-baseweb="tab-list"] {{
    background: {DARK}; border-radius: 10px;
    border: 1px solid {BORDER}; padding: 4px; gap: 4px;
  }}
  .stTabs [data-baseweb="tab"] {{
    color: {LGRAY}; border-radius: 7px; font-weight: 600;
    font-size: 0.82rem; letter-spacing: 0.5px;
  }}
  .stTabs [aria-selected="true"] {{
    background: linear-gradient(135deg, {GOLD}22, {TEAL}22) !important;
    color: {GOLD} !important;
    border-bottom: 2px solid {GOLD} !important;
  }}

  /* Buttons */
  .stButton > button {{
    background: linear-gradient(135deg, {GOLD}, #D09020);
    color: {BLACK}; border: none; border-radius: 8px;
    font-weight: 700; letter-spacing: 0.5px;
  }}
  .stDownloadButton > button {{
    background: linear-gradient(135deg, {TEAL}, #1A6060);
    color: {WHITE}; border: none; border-radius: 8px;
    font-weight: 700; width: 100%;
  }}

  /* Filter tag pills */
  .filter-pill {{
    display: inline-block;
    background: {GOLD}22; border: 1px solid {GOLD}55;
    color: {GOLD}; border-radius: 20px;
    padding: 2px 10px; font-size: 0.7rem; font-weight: 600;
    margin: 2px;
  }}

  /* Sidebar label */
  .sb-label {{
    font-size: 0.7rem; font-weight: 700; letter-spacing: 1.5px;
    color: {GOLD}; text-transform: uppercase; margin-bottom: 4px;
    display: block;
  }}

  /* Dataframe styling */
  .stDataFrame {{ border-radius: 10px; overflow: hidden; }}
  iframe {{ border-radius: 10px; }}

  /* Divider */
  hr {{ border-color: {BORDER}; margin: 16px 0; }}

  /* Insight box */
  .insight-box {{
    background: {DARK}; border: 1px solid {BORDER};
    border-radius: 10px; padding: 14px 16px; margin-top: 10px;
  }}
  .insight-text {{ font-size: 0.82rem; color: {LGRAY}; line-height: 1.6; }}
  .insight-highlight {{ color: {GOLD}; font-weight: 700; }}
</style>
""", unsafe_allow_html=True)

# ── HEADER ────────────────────────────────────────────────────
logo_html = f'<img src="data:image/png;base64,{logo_b64}" height="48" style="filter:brightness(10);">' if logo_b64 else '📊'
st.markdown(f"""
<div class="me-header">
  <div class="me-brand">
    {logo_html}
    <div>
      <div class="me-title">MARIANO EVENTS</div>
      <div class="me-subtitle">CAMPAIGN INTELLIGENCE PLATFORM</div>
    </div>
  </div>
  <div style="text-align:right;">
    <div style="font-size:0.7rem;color:#555;letter-spacing:1px;">POWERED BY</div>
    <div style="font-size:0.85rem;font-weight:700;color:{GOLD};">PROMOMASH ANALYTICS</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── CHART HELPERS ─────────────────────────────────────────────
def setup_fig(nrows=1, ncols=1, figw=14, figh=4):
    fig, axes = plt.subplots(nrows, ncols, figsize=(figw, figh))
    fig.patch.set_facecolor('#111111')
    if nrows==1 and ncols==1: axes=[axes]
    axes_flat = axes if isinstance(axes, list) else \
                [axes] if not hasattr(axes,'flat') else list(axes.flat)
    for ax in axes_flat:
        ax.set_facecolor('#1A1A1A')
        ax.tick_params(colors='#888888', labelsize=8.5)
        ax.xaxis.label.set_color('#888888')
        ax.yaxis.label.set_color('#888888')
        ax.title.set_color(GOLD)
        ax.title.set_fontsize(11)
        ax.title.set_fontweight('bold')
        for spine in ax.spines.values():
            spine.set_edgecolor('#2A2A2A')
        ax.grid(axis='y', color='#2A2A2A', linewidth=0.5, linestyle='--')
        ax.grid(axis='x', visible=False)
        ax.set_axisbelow(True)
    return fig, axes_flat

def hsetup(ax):
    ax.grid(axis='x', color='#2A2A2A', linewidth=0.5, linestyle='--')
    ax.grid(axis='y', visible=False)

def blbl(ax, bars, fmt='{:.0f}', horiz=False, color=WHITE, fs=9):
    for bar in bars:
        v = bar.get_width() if horiz else bar.get_height()
        if v > 0:
            if horiz:
                ax.text(v*1.01+0.05, bar.get_y()+bar.get_height()/2,
                        fmt.format(v), va='center', color=color, fontsize=fs, fontweight='bold')
            else:
                ax.text(bar.get_x()+bar.get_width()/2, v*1.01+0.05,
                        fmt.format(v), ha='center', va='bottom', color=color, fontsize=fs, fontweight='bold')

def gradient_bar(ax, x, y, color1, color2, width=0.6):
    from matplotlib.patches import Rectangle
    from matplotlib.collections import PatchCollection
    n_steps = 50
    for i, (xi, yi) in enumerate(zip(x, y)):
        if yi <= 0: continue
        for j in range(n_steps):
            frac = j/n_steps
            r = int(int(color1[1:3],16)*(1-frac) + int(color2[1:3],16)*frac)
            g = int(int(color1[3:5],16)*(1-frac) + int(color2[3:5],16)*frac)
            b = int(int(color1[5:7],16)*(1-frac) + int(color2[5:7],16)*frac)
            c = f'#{r:02x}{g:02x}{b:02x}'
            rect = Rectangle((i-width/2, yi*j/n_steps), width, yi/n_steps, color=c)
            ax.add_patch(rect)
    ax.set_xlim(-0.5, len(x)-0.5)
    ax.set_ylim(0, max(y)*1.15)
    ax.set_xticks(range(len(x)))
    ax.set_xticklabels(x, fontsize=9)

def chart_buf(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=130, bbox_inches='tight',
                facecolor='#111111', edgecolor='none')
    buf.seek(0); plt.close(fig); return buf

# ── DATA HELPERS ──────────────────────────────────────────────
def parse_coupons(x):
    if pd.isna(x): return 0
    nums = re.findall(r'\d+', str(x))
    return int(nums[0]) if nums else 0

def best_match(cols, hints):
    cl = [c.lower().strip() for c in cols]
    for h in hints:
        for i,c in enumerate(cl):
            if h.lower() in c: return cols[i]
    return None

def guess_mapping(cols):
    return {
        'date':        best_match(cols, ['date','event date']),
        'campaign':    best_match(cols, ['campaign name','campaign','program']),
        'ba_name':     best_match(cols, ['ba name','ambassador','rep name','staff name']),
        'venue':       best_match(cols, ['venue name','store name','location name']),
        'city':        best_match(cols, ['city']),
        'state':       best_match(cols, ['state']),
        'units':       best_match(cols, ['# total sales','total sales','units sold','# of units']),
        'revenue':     best_match(cols, ['$ total sales','revenue','dollar']),
        'sampled':     best_match(cols, ['total sampled','sampled','samples given']),
        'hours':       best_match(cols, ['ba hours','hours worked','hours']),
        'coupons':     best_match(cols, ['coupon']),
        'sku1':        best_match(cols, ['pineapple','pnapple','guava']),
        'sku2':        best_match(cols, ['lemon lime','lemon']),
        'sku3':        best_match(cols, ['orange mango','mango']),
        'shelf':       best_match(cols, ['enough product','shelf','stock support']),
        'inventory':   best_match(cols, ['inventory concern','inventory']),
        'impact':      best_match(cols, ['impacted','impact','affect']),
        'comments':    best_match(cols, ['customer comments','comments','feedback']),
        'familiarity': best_match(cols, ['familiar','brand awareness']),
        'recommend':   best_match(cols, ['recommend','likelihood']),
        'taste':       best_match(cols, ['taste','flavor rating']),
    }

def prepare(df, m):
    r = df.copy()
    r['_units']   = pd.to_numeric(r[m['units']], errors='coerce').fillna(0)
    r['_sampled'] = pd.to_numeric(r[m['sampled']], errors='coerce').fillna(0)
    r['_revenue'] = pd.to_numeric(r[m['revenue']], errors='coerce').fillna(0) if m.get('revenue') else 0
    r['_hours']   = pd.to_numeric(r[m['hours']], errors='coerce').fillna(0) if m.get('hours') else 0
    r['_coupons'] = r[m['coupons']].apply(parse_coupons) if m.get('coupons') else 0
    r['_date']    = pd.to_datetime(r[m['date']], errors='coerce')
    r['_dow']     = r['_date'].dt.day_name()
    r['_conv']    = (r['_units'] / r['_sampled'].replace(0,np.nan) * 100).round(1)
    for k in ['sku1','sku2','sku3']:
        r[f'_{k}'] = pd.to_numeric(r[m[k]], errors='coerce').fillna(0) if m.get(k) else 0
    r['_campaign'] = r[m['campaign']].fillna('Unknown').astype(str) if m.get('campaign') else 'Campaign'
    r['_state']    = r[m['state']].fillna('Unknown').astype(str) if m.get('state') else 'Unknown'
    r['_city']     = r[m['city']].fillna('Unknown').astype(str) if m.get('city') else 'Unknown'
    r['_ba']       = r[m['ba_name']].fillna('Unknown').astype(str) if m.get('ba_name') else 'Unknown'
    r['_venue']    = r[m['venue']].fillna('Unknown').astype(str) if m.get('venue') else 'Unknown'

    # Stock analysis
    def blob(row):
        cols_check = ['shelf','inventory','impact','comments']
        return ' '.join(str(row[m[c]]) for c in cols_check if m.get(c) and c in row.index and pd.notna(row.get(m[c],''))).lower()
    r['_blob'] = r.apply(blob, axis=1)
    r['_sold_out']  = r['_blob'].str.contains('sold out|ran out|out of stock|selling out|sold through', na=False)
    r['_low_stock'] = r['_blob'].str.contains('low stock|low inventory|restock|not enough|limited stock|running low', na=False)
    r['_no_shelf']  = (r[m['shelf']] == 'No') if m.get('shelf') else False
    r['_stock_issue'] = r['_sold_out'] | r['_low_stock'] | r['_no_shelf']
    return r

def compute_kpis(r):
    n=len(r); units=r['_units'].sum(); samp=r['_sampled'].sum()
    rev=r['_revenue'].sum(); hrs=r['_hours'].sum()
    conv=round(units/samp*100,2) if samp>0 else 0
    avg_u=round(units/n,2) if n>0 else 0; avg_s=round(samp/n,1) if n>0 else 0
    uph=round(units/hrs,2) if hrs>0 else 0; avg_r=round(rev/n,2) if n>0 else 0
    coup=r['_coupons'].sum(); coup_e=(r['_coupons']>0).sum()
    wc=r[r['_coupons']>0]['_units'].mean() if (r['_coupons']>0).any() else 0
    nc=r[r['_coupons']==0]['_units'].mean() if (r['_coupons']==0).any() else 0
    s1=r['_sku1'].sum(); s2=r['_sku2'].sum(); s3=r['_sku3'].sum()
    stock_e=r['_stock_issue'].sum(); soldout=r['_sold_out'].sum()
    return dict(n=int(n),units=float(units),samp=float(samp),rev=float(rev),hrs=float(hrs),
                conv=conv,avg_u=avg_u,avg_s=avg_s,uph=uph,avg_r=avg_r,
                coup=float(coup),coup_e=int(coup_e),wc=round(wc,2),nc=round(nc,2),
                s1=float(s1),s2=float(s2),s3=float(s3),
                stock_e=int(stock_e),soldout=int(soldout))

def sec(title, color='gold', icon='📊'):
    color_map = {'gold':('sec-head','sec-title'),'purple':('sec-head purple','sec-title purple'),
                 'teal':('sec-head teal','sec-title teal'),'green':('sec-head green','sec-title green'),
                 'red':('sec-head red','sec-title red')}
    hc,tc = color_map.get(color,('sec-head','sec-title'))
    st.markdown(f'<div class="{hc}"><span class="{tc}">{icon} {title}</span></div>', unsafe_allow_html=True)

def kpi_row(cards):
    cols = st.columns(len(cards))
    for col,(val,lbl,color,sub) in zip(cols,cards):
        cls = f'kpi-card {color}'
        vcls = f'kpi-val {color}'
        sub_html = f'<div class="kpi-sub">{sub}</div>' if sub else ''
        col.markdown(f'<div class="{cls}"><div class="{vcls}">{val}</div><div class="kpi-lbl">{lbl}</div>{sub_html}</div>',
                     unsafe_allow_html=True)

# ── SIDEBAR ────────────────────────────────────────────────────
with st.sidebar:
    if logo_b64:
        st.markdown(f'<div style="text-align:center;padding:12px 0 6px 0;"><img src="data:image/png;base64,{logo_b64}" height="60" style="filter:brightness(10);"></div>', unsafe_allow_html=True)
    st.markdown(f'<div style="text-align:center;font-size:0.65rem;color:#555;letter-spacing:2px;margin-bottom:16px;">CAMPAIGN ANALYZER</div>', unsafe_allow_html=True)
    st.markdown(f'<div style="height:1px;background:{BORDER};margin-bottom:16px;"></div>', unsafe_allow_html=True)

    st.markdown('<span class="sb-label">📁 Upload Data</span>', unsafe_allow_html=True)
    uploaded_files = st.file_uploader(
        "", type=['xlsx','xls','csv'],
        accept_multiple_files=True,
        help="Upload one or more Promomash Excel files"
    )
    if not uploaded_files:
        st.markdown(f'<div style="background:{DARK};border:1px dashed {BORDER};border-radius:10px;padding:20px;text-align:center;margin-top:8px;"><div style="color:{GOLD};font-size:1.5rem;">📂</div><div style="color:{LGRAY};font-size:0.8rem;margin-top:8px;">Upload your Promomash file to get started</div></div>', unsafe_allow_html=True)
        st.stop()

# ── LOAD DATA ─────────────────────────────────────────────────
@st.cache_data
def load_file(fname, data):
    try:
        xl = pd.ExcelFile(io.BytesIO(data))
        df = xl.parse('Raw') if 'Raw' in xl.sheet_names else xl.parse(xl.sheet_names[0])
    except:
        df = pd.read_csv(io.BytesIO(data))
    return df[df.iloc[:,0].notna()].copy()

all_dfs = {}
for uf in uploaded_files:
    df = load_file(uf.name, uf.read())
    cname = uf.name.replace('.xlsx','').replace('.xls','').replace('.csv','')
    all_dfs[cname] = df

combined_raw = pd.concat(all_dfs.values(), ignore_index=True)
cols = combined_raw.columns.tolist()

# ── COLUMN MAPPING ─────────────────────────────────────────────
if 'mapping' not in st.session_state:
    st.session_state.mapping = guess_mapping(cols)
m = st.session_state.mapping

required = ['date','ba_name','units','sampled']
missing = [k for k in required if not m.get(k)]

with st.sidebar:
    st.markdown(f'<div style="height:1px;background:{BORDER};margin:12px 0;"></div>', unsafe_allow_html=True)
    if st.checkbox("⚙️ Edit Column Mapping", value=bool(missing)):
        st.markdown('<span class="sb-label">Column Mapping</span>', unsafe_allow_html=True)
        none_opt = '— Not in file —'
        all_opts = [none_opt] + cols
        new_m = {}
        FIELDS = [
            ('date','Date ✱',True),('ba_name','BA Name ✱',True),
            ('units','Units Sold ✱',True),('sampled','Total Sampled ✱',True),
            ('revenue','Revenue ($)',False),('hours','BA Hours',False),
            ('coupons','Coupons',False),('campaign','Campaign',False),
            ('venue','Store / Venue',False),('city','City',False),
            ('state','State',False),('sku1','SKU - Pineapple Guava',False),
            ('sku2','SKU - Lemon Lime',False),('sku3','SKU - Orange Mango',False),
            ('shelf','Enough Product on Shelf?',False),
            ('inventory','Inventory Concerns',False),
            ('impact','What Impacted Sales?',False),
            ('comments','Customer Comments',False),
        ]
        for key,label,req in FIELDS:
            guess = m.get(key)
            idx = all_opts.index(guess) if guess and guess in all_opts else 0
            sel = st.selectbox(label, all_opts, index=idx, key=f'col_{key}')
            new_m[key] = None if sel==none_opt else sel
        if st.button("✅ Save Mapping"):
            st.session_state.mapping = new_m; m = new_m; st.rerun()

    st.markdown(f'<div style="height:1px;background:{BORDER};margin:12px 0;"></div>', unsafe_allow_html=True)

# ── PREPARE DATA ───────────────────────────────────────────────
r_full = combined_raw[pd.to_numeric(combined_raw[m['units']], errors='coerce').notna()].copy()
r_full = prepare(r_full, m)

# ── SIDEBAR FILTERS ────────────────────────────────────────────
with st.sidebar:
    st.markdown('<span class="sb-label">🎯 Filters</span>', unsafe_allow_html=True)

    camps = sorted(r_full['_campaign'].unique())
    sel_c = st.multiselect("Campaign", camps, default=camps)

    min_d = r_full['_date'].min().date()
    max_d = r_full['_date'].max().date()
    dr = st.date_input("Date Range", value=(min_d, max_d), min_value=min_d, max_value=max_d)

    states = sorted(r_full['_state'].unique())
    sel_s = st.multiselect("State", states, default=states)

    bas = sorted(r_full['_ba'].unique())
    sel_b = st.multiselect("BA / Ambassador", bas, default=bas)

    st.markdown(f'<div style="height:1px;background:{BORDER};margin:12px 0;"></div>', unsafe_allow_html=True)
    st.markdown('<span class="sb-label">⚙️ Display Settings</span>', unsafe_allow_html=True)
    top_n = st.slider("Top / Bottom N", 5, 20, 10)
    min_ev = st.slider("Min events for BA rank", 1, 5, 2)

# ── APPLY FILTERS ──────────────────────────────────────────────
r = r_full[r_full['_campaign'].isin(sel_c)].copy()
if len(dr)==2:
    r = r[(r['_date'].dt.date>=dr[0])&(r['_date'].dt.date<=dr[1])]
r = r[r['_state'].isin(sel_s) & r['_ba'].isin(sel_b)]

if len(r)==0:
    st.warning("No data matches the current filters.")
    st.stop()

k = compute_kpis(r)

# ── KPI BANNER ────────────────────────────────────────────────
st.markdown(f'<div style="color:{LGRAY};font-size:0.78rem;margin-bottom:10px;">Showing <span style="color:{GOLD};font-weight:700;">{k["n"]:,} events</span> · {dr[0] if len(dr)>0 else ""} to {dr[1] if len(dr)>1 else ""} · {len(sel_c)} campaign(s) · {len(sel_s)} state(s)</div>', unsafe_allow_html=True)

kpi_row([
    (f"{k['n']:,}", "TOTAL EVENTS", "", None),
    (f"{k['samp']:,.0f}", "SAMPLES GIVEN", "teal", None),
    (f"{k['units']:,.0f}", "UNITS SOLD (BA)", "purple", None),
    (f"${k['rev']:,.0f}", "REVENUE", "", None),
    (f"{k['conv']}%", "CONVERSION RATE", "teal", None),
    (f"{k['avg_u']}", "AVG UNITS/DEMO", "", None),
    (f"{k['avg_s']}", "AVG SAMPLED/DEMO", "teal", None),
    (f"{k['uph']}", "UNITS/BA HOUR", "purple", None),
])
st.markdown("<br>", unsafe_allow_html=True)

coupon_lift = round((k['wc']/k['nc']-1)*100,1) if k['nc']>0 else 0
stock_pct = round(k['stock_e']/k['n']*100,1) if k['n']>0 else 0
kpi_row([
    (f"{k['coup_e']}", "EVENTS WITH COUPONS", "gold", f"of {k['n']} total"),
    (f"+{coupon_lift}%", "COUPON LIFT", "green", f"{k['wc']:.1f} vs {k['nc']:.1f} avg units"),
    (f"{k['stock_e']}", "STOCK ISSUE EVENTS", "red", f"{stock_pct}% of all events"),
    (f"{k['soldout']}", "SELL-OUTS", "red", "demand signal not a failure"),
])

st.markdown(f'<div style="height:1px;background:{BORDER};margin:16px 0;"></div>', unsafe_allow_html=True)

# ── MAIN TABS ─────────────────────────────────────────────────
tabs = st.tabs(["📅 Trends", "🏆 BA Performance", "🏪 Stores",
                "🗺️ Geography", "🎯 Sampling & Coupons",
                "📦 Stock & Inventory", "🥤 SKUs",
                "📊 Campaign Compare", "📥 Download"])

charts_excel = {}

# ── TAB 1: TRENDS ─────────────────────────────────────────────
with tabs[0]:
    daily = r.groupby('_date').agg(n=('_units','size'),u=('_units','sum'),
                                    s=('_sampled','sum')).reset_index()
    daily['avg']=(daily.u/daily.n).round(1)
    daily['conv']=(daily.u/daily.s*100).round(1)

    sec("Daily Performance Trend", "gold", "📅")
    fig, axs = setup_fig(1,2,14,4)
    x = range(len(daily))
    axs[0].fill_between(x, daily['u'], alpha=0.3, color=GOLD)
    axs[0].plot(x, daily['u'], color=GOLD, linewidth=2.5, marker='o', markersize=4)
    axs[0].set_xticks(x[::max(1,len(x)//8)])
    axs[0].set_xticklabels([str(d)[:10] for d in daily['_date'].iloc[::max(1,len(x)//8)]], rotation=35, ha='right', fontsize=7.5)
    axs[0].set_title('Total Units Sold per Day', pad=10)
    axs[0].set_ylabel('Units Sold', color=LGRAY)

    axs[1].fill_between(x, daily['avg'], alpha=0.3, color=TEAL)
    axs[1].plot(x, daily['avg'], color=TEAL, linewidth=2.5, marker='o', markersize=4)
    axs[1].set_xticks(x[::max(1,len(x)//8)])
    axs[1].set_xticklabels([str(d)[:10] for d in daily['_date'].iloc[::max(1,len(x)//8)]], rotation=35, ha='right', fontsize=7.5)
    axs[1].set_title('Avg Units per Demo per Day', pad=10)
    axs[1].set_ylabel('Avg Units', color=LGRAY)
    plt.tight_layout(); st.pyplot(fig)
    charts_excel['Daily Trend'] = chart_buf(fig)

    sec("Day of Week Performance", "teal", "📆")
    order=['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
    dow = r.groupby('_dow').agg(n=('_units','size'),u=('_units','sum'),s=('_sampled','sum')).reset_index()
    dow['avg']=(dow.u/dow.n).round(2); dow['conv']=(dow.u/dow.s*100).round(2)
    dow['sort']=dow['_dow'].map({d:i for i,d in enumerate(order)})
    dow=dow.sort_values('sort')

    fig, axs = setup_fig(1,2,14,4)
    colors=[GOLD if d in ['Friday','Saturday','Sunday'] else TEAL for d in dow['_dow']]
    bars=axs[0].bar(dow['_dow'],dow['avg'],color=colors,edgecolor='none',width=0.65)
    axs[0].set_title('Avg Units/Demo by Day', pad=10)
    blbl(axs[0],bars,'{:.1f}',fs=9)
    plt.setp(axs[0].xaxis.get_majorticklabels(),rotation=25,ha='right')

    bars=axs[1].bar(dow['_dow'],dow['conv'],color=colors,edgecolor='none',width=0.65,alpha=0.85)
    axs[1].set_title('Conversion % by Day', pad=10)
    blbl(axs[1],bars,'{:.1f}%',fs=9)
    plt.setp(axs[1].xaxis.get_majorticklabels(),rotation=25,ha='right')
    plt.tight_layout(); st.pyplot(fig)
    charts_excel['By Day'] = chart_buf(fig)

    st.dataframe(dow[['_dow','n','u','avg','conv']].rename(columns={
        '_dow':'Day','n':'Events','u':'Units','avg':'Avg Units/Demo','conv':'Conv %'
    }), use_container_width=True, hide_index=True)

# ── TAB 2: BA PERFORMANCE ─────────────────────────────────────
with tabs[1]:
    ba=r.groupby('_ba').agg(n=('_units','size'),u=('_units','sum'),
        rev=('_revenue','sum'),s=('_sampled','sum')).reset_index()
    ba['avg']=(ba.u/ba.n).round(2); ba['conv']=(ba.u/ba.s*100).round(2)
    ba_q=ba[ba.n>=min_ev]
    top_ba=ba_q.sort_values('avg',ascending=False).head(top_n)
    low_ba=ba_q.sort_values('avg').head(top_n)

    sec(f"Top {top_n} Brand Ambassadors", "gold", "🥇")
    fig, axs = setup_fig(1,2,14,max(4,top_n*0.5))
    hsetup(axs[0]); hsetup(axs[1])
    top_colors = [GOLD if i==0 else TEAL if i==1 else '#888' for i in range(len(top_ba))]
    bars=axs[0].barh(top_ba['_ba'],top_ba['avg'],color=top_colors,edgecolor='none')
    axs[0].set_title(f'Top {top_n} BAs — Avg Units/Demo',pad=10)
    axs[0].set_xlabel('Avg Units / Demo',color=LGRAY); axs[0].invert_yaxis()
    blbl(axs[0],bars,'{:.2f}',horiz=True,color=GOLD)

    bars=axs[1].barh(low_ba['_ba'],low_ba['avg'],color=RED,alpha=0.8,edgecolor='none')
    axs[1].set_title(f'Bottom {top_n} BAs — Avg Units/Demo',pad=10)
    axs[1].set_xlabel('Avg Units / Demo',color=LGRAY); axs[1].invert_yaxis()
    blbl(axs[1],bars,'{:.2f}',horiz=True,color='#FF9999')
    plt.tight_layout(); st.pyplot(fig)
    charts_excel['By BA'] = chart_buf(fig)

    col1,col2=st.columns(2)
    with col1:
        st.markdown(f'<div style="color:{GOLD};font-size:0.8rem;font-weight:700;margin-bottom:6px;">TOP {top_n} BAS</div>',unsafe_allow_html=True)
        st.dataframe(top_ba[['_ba','n','u','avg','conv']].rename(columns={
            '_ba':'BA Name','n':'Events','u':'Units','avg':'Avg Units/Demo','conv':'Conv %'}),
            use_container_width=True, hide_index=True)
    with col2:
        st.markdown(f'<div style="color:{RED};font-size:0.8rem;font-weight:700;margin-bottom:6px;">BOTTOM {top_n} BAS</div>',unsafe_allow_html=True)
        st.dataframe(low_ba[['_ba','n','u','avg','conv']].rename(columns={
            '_ba':'BA Name','n':'Events','u':'Units','avg':'Avg Units/Demo','conv':'Conv %'}),
            use_container_width=True, hide_index=True)

# ── TAB 3: STORES ─────────────────────────────────────────────
with tabs[2]:
    st_df=r.groupby(['_venue','_city','_state']).agg(
        n=('_units','size'),u=('_units','sum'),
        rev=('_revenue','sum'),s=('_sampled','sum')).reset_index()
    st_df['avg']=(st_df.u/st_df.n).round(2)
    st_df['conv']=(st_df.u/st_df.s*100).round(2)
    top_st=st_df.sort_values('u',ascending=False).head(top_n).copy()
    low_st=st_df.sort_values('u').head(top_n).copy()
    top_st['lbl']=top_st['_city']+', '+top_st['_state']
    low_st['lbl']=low_st['_city']+', '+low_st['_state']

    sec(f"Top & Bottom {top_n} Stores", "gold", "🏪")
    fig, axs = setup_fig(1,2,14,max(4,top_n*0.5))
    hsetup(axs[0]); hsetup(axs[1])
    tc=[GOLD if i==0 else TEAL if i==1 else PURPLE for i in range(len(top_st))]
    bars=axs[0].barh(top_st['lbl'],top_st['u'],color=tc,edgecolor='none')
    axs[0].set_title(f'Top {top_n} Stores — Total Units',pad=10); axs[0].invert_yaxis()
    blbl(axs[0],bars,'{:.0f}',horiz=True,color=GOLD)

    bars=axs[1].barh(low_st['lbl'],low_st['u'],color=RED,alpha=0.8,edgecolor='none')
    axs[1].set_title(f'Bottom {top_n} Stores — Total Units',pad=10); axs[1].invert_yaxis()
    blbl(axs[1],bars,'{:.0f}',horiz=True,color='#FF9999')
    plt.tight_layout(); st.pyplot(fig)
    charts_excel['Stores'] = chart_buf(fig)

    col1,col2=st.columns(2)
    with col1:
        st.dataframe(top_st[['lbl','_venue','n','u','avg','conv']].rename(columns={
            'lbl':'Location','_venue':'Store','n':'Events','u':'Units','avg':'Avg/Demo','conv':'Conv %'}),
            use_container_width=True, hide_index=True)
    with col2:
        st.dataframe(low_st[['lbl','_venue','n','u','avg','conv']].rename(columns={
            'lbl':'Location','_venue':'Store','n':'Events','u':'Units','avg':'Avg/Demo','conv':'Conv %'}),
            use_container_width=True, hide_index=True)

# ── TAB 4: GEOGRAPHY ──────────────────────────────────────────
with tabs[3]:
    sg=r.groupby('_state').agg(n=('_units','size'),u=('_units','sum'),
        s=('_sampled','sum'),rev=('_revenue','sum')).reset_index()
    sg['avg']=(sg.u/sg.n).round(2); sg['conv']=(sg.u/sg.s*100).round(2)
    sg=sg.sort_values('u',ascending=False)
    avg_line=sg['avg'].mean()

    sec("State Performance", "teal", "🗺️")
    fig, axs = setup_fig(1,2,14,4)
    sc=[GREEN if v>=avg_line else RED for v in sg['avg']]
    bars=axs[0].bar(sg['_state'],sg['avg'],color=sc,edgecolor='none',width=0.7)
    axs[0].axhline(avg_line,color=GOLD,linewidth=1.5,linestyle='--',label=f'Avg {avg_line:.1f}')
    axs[0].legend(facecolor=DARK,labelcolor=WHITE,fontsize=8)
    axs[0].set_title('Avg Units/Demo by State',pad=10)
    plt.setp(axs[0].xaxis.get_majorticklabels(),rotation=45,ha='right',fontsize=8)

    bars=axs[1].bar(sg['_state'],sg['n'],color=TEAL,edgecolor='none',width=0.7,alpha=0.85)
    axs[1].set_title('Events by State',pad=10)
    blbl(axs[1],bars,'{:.0f}',fs=8)
    plt.setp(axs[1].xaxis.get_majorticklabels(),rotation=45,ha='right',fontsize=8)
    plt.tight_layout(); st.pyplot(fig)
    charts_excel['By State'] = chart_buf(fig)

    st.dataframe(sg.rename(columns={'_state':'State','n':'Events','u':'Units',
        's':'Sampled','rev':'Revenue ($)','avg':'Avg/Demo','conv':'Conv %'}),
        use_container_width=True, hide_index=True)

# ── TAB 5: SAMPLING & COUPONS ─────────────────────────────────
with tabs[4]:
    r2=r.copy()
    r2['tier']=pd.cut(r2['_sampled'],bins=[0,50,65,80,300],labels=['≤50','51-65','66-80','80+'])
    sg2=r2.groupby('tier',observed=True).agg(n=('_units','size'),u=('_units','sum'),
        s=('_sampled','sum')).reset_index()
    sg2['avg']=(sg2.u/sg2.n).round(2)

    sec("Sampling Volume Impact", "teal", "🎯")
    fig, axs = setup_fig(1,2,12,4)
    tc2=[RED,GOLD,TEAL,GREEN]
    bars=axs[0].bar(sg2['tier'].astype(str),sg2['avg'],color=tc2,edgecolor='none',width=0.65)
    axs[0].set_title('Avg Units/Demo by Sampling Volume',pad=10)
    axs[0].set_xlabel('Customers Sampled',color=LGRAY)
    blbl(axs[0],bars,'{:.2f}',fs=10,color=WHITE)

    axs[1].set_facecolor('#111111')
    wedges,texts,autos=axs[1].pie(sg2['n'],labels=sg2['tier'].astype(str),colors=tc2,
        autopct='%1.0f%%',startangle=90,textprops={'color':WHITE,'fontsize':10})
    axs[1].set_title('Event Distribution by Tier',pad=10)
    plt.tight_layout(); st.pyplot(fig)
    charts_excel['Sampling'] = chart_buf(fig)

    sec("Coupon Impact Analysis", "gold", "🎟️")
    wc=r[r['_coupons']>0]['_units'].mean() if (r['_coupons']>0).any() else 0
    nc=r[r['_coupons']==0]['_units'].mean() if (r['_coupons']==0).any() else 0
    lift=round((wc/nc-1)*100,1) if nc>0 else 0

    fig, axs = setup_fig(1,2,12,4)
    bars=axs[0].bar(['With Coupons','Without Coupons'],[wc,nc],
                    color=[GREEN,RED],edgecolor='none',width=0.5)
    axs[0].set_title('Coupon Impact on Avg Units/Demo',pad=10)
    blbl(axs[0],bars,'{:.2f}',fs=12,color=WHITE)
    axs[0].text(0.5,0.93,f'Lift: +{lift}%',transform=axs[0].transAxes,
                ha='center',color=GOLD,fontsize=12,fontweight='bold')

    coup_daily=r.groupby('_date').apply(lambda g: pd.Series({'with':g[g['_coupons']>0]['_units'].mean() if (g['_coupons']>0).any() else 0,
        'without':g[g['_coupons']==0]['_units'].mean() if (g['_coupons']==0).any() else 0})).reset_index()
    x=range(len(coup_daily))
    axs[1].fill_between(x,coup_daily['with'],alpha=0.25,color=GREEN)
    axs[1].plot(x,coup_daily['with'],color=GREEN,linewidth=2,label='With coupons')
    axs[1].fill_between(x,coup_daily['without'],alpha=0.25,color=RED)
    axs[1].plot(x,coup_daily['without'],color=RED,linewidth=2,label='Without coupons')
    axs[1].legend(facecolor=DARK,labelcolor=WHITE,fontsize=8)
    axs[1].set_title('Daily Units: Coupon vs No Coupon',pad=10)
    plt.tight_layout(); st.pyplot(fig)
    charts_excel['Coupons'] = chart_buf(fig)

# ── TAB 6: STOCK & INVENTORY ──────────────────────────────────
with tabs[5]:
    sec("Stock & Inventory Analysis", "red", "📦")

    # KPIs
    stock_ev=r[r['_stock_issue']]
    no_stock=r[~r['_stock_issue']]
    avg_with=stock_ev['_units'].mean() if len(stock_ev)>0 else 0
    avg_without=no_stock['_units'].mean() if len(no_stock)>0 else 0

    kpi_row([
        (f"{k['stock_e']}", "EVENTS WITH STOCK ISSUE", "red", f"{stock_pct}% of all events"),
        (f"{k['soldout']}", "CONFIRMED SELL-OUTS", "red", "ran out during demo"),
        (f"{avg_with:.1f}", "AVG UNITS W/ STOCK ISSUE", "gold", "compare to baseline below"),
        (f"{avg_without:.1f}", "AVG UNITS — FULL SHELVES", "teal", "baseline performance"),
    ])

    # Insight box
    diff = avg_with - avg_without
    direction = "MORE" if diff > 0 else "FEWER"
    color_i = GREEN if diff > 0 else RED
    st.markdown(f"""
    <div class="insight-box" style="border-color:{color_i}44;border-left:3px solid {color_i};">
      <div class="insight-text">
        Events that reported a stock issue sold <span class="insight-highlight" style="color:{color_i};">{abs(diff):.2f} units {direction}</span>
        than events with full shelves.
        {"This means stores are selling out BECAUSE the product moves fast — it is a demand signal, not a failure." if diff>0 else "Stock issues are hurting sales at these locations. Prioritize restocking before the next demo."}
      </div>
    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Chart
    fig, axs = setup_fig(1,2,12,4)
    bars=axs[0].bar(['Stock Issue Events','Full Shelf Events'],[avg_with,avg_without],
                    color=[GOLD,TEAL],edgecolor='none',width=0.5)
    axs[0].set_title('Avg Units: Stock Issue vs Full Shelf',pad=10)
    blbl(axs[0],bars,'{:.2f}',fs=12,color=WHITE)

    # Stock issues by state
    if m.get('state'):
        sg_s=r.groupby('_state').agg(
            total=('_stock_issue','count'), issues=('_stock_issue','sum')).reset_index()
        sg_s['rate']=(sg_s.issues/sg_s.total*100).round(1)
        sg_s=sg_s[sg_s.total>=2].sort_values('rate',ascending=False).head(12)
        sc3=[RED if v>30 else GOLD if v>15 else TEAL for v in sg_s['rate']]
        bars=axs[1].bar(sg_s['_state'],sg_s['rate'],color=sc3,edgecolor='none',width=0.7)
        axs[1].set_title('Stock Issue Rate by State (%)',pad=10)
        blbl(axs[1],bars,'{:.0f}%',fs=8,color=WHITE)
        plt.setp(axs[1].xaxis.get_majorticklabels(),rotation=45,ha='right',fontsize=8)
    plt.tight_layout(); st.pyplot(fig)
    charts_excel['Stock'] = chart_buf(fig)

    # Stock issue event list
    col1,col2=st.columns(2)
    with col1:
        sec("Sell-Out Events", "red", "🔴")
        so=r[r['_sold_out']][['_ba','_city','_state','_units','_sampled']].copy()
        if len(so)>0:
            so.columns=['BA','City','State','Units','Sampled']
            so['Conv%']=(so['Units']/so['Sampled']*100).round(1)
            st.dataframe(so.sort_values('Units',ascending=False),use_container_width=True,hide_index=True)
            st.markdown(f'<div class="stock-card ok"><div class="stock-title">✅ Sell-outs sold {(avg_with/avg_without-1)*100:.0f}% more than full shelf events</div><div class="stock-detail">These stores ran out because the product moves fast. They need larger opening orders.</div></div>',unsafe_allow_html=True)
        else:
            st.info("No sell-out events in current filter.")

    with col2:
        sec("Low Stock Events", "gold", "🟡")
        ls=r[r['_low_stock']][['_ba','_city','_state','_units']].copy()
        if len(ls)>0:
            ls.columns=['BA','City','State','Units']
            st.dataframe(ls.sort_values('Units',ascending=False),use_container_width=True,hide_index=True)
            st.markdown(f'<div class="stock-card warn"><div class="stock-title">⚠️ {len(ls)} events reported low stock during the demo</div><div class="stock-detail">BAs flagged running low mid-event. Pre-event inventory checks can prevent this.</div></div>',unsafe_allow_html=True)
        else:
            st.info("No low stock events in current filter.")

    # Field quotes
    if m.get('impact') or m.get('inventory') or m.get('comments'):
        sec("Field Comments — Stock Related", "gold", "💬")
        def get_stock_comments(r):
            rows=[]
            for col in ['impact','inventory','comments']:
                if m.get(col):
                    sub=r[r['_stock_issue']][[m[col],'_city','_state']].dropna(subset=[m[col]])
                    sub=sub[sub[m[col]].astype(str).str.lower().str.contains(
                        'stock|sold out|inventory|shelf|ran out|low|restock', na=False)]
                    for _,row in sub.iterrows():
                        txt=str(row[m[col]]).strip()
                        if 10<len(txt)<300 and txt.lower() not in ['none','n/a','na','no']:
                            rows.append({'Location':f"{row['_city']}, {row['_state']}",'Comment':txt})
            return pd.DataFrame(rows).drop_duplicates('Comment').head(10)
        cmt=get_stock_comments(r)
        if len(cmt)>0:
            for _,row in cmt.iterrows():
                st.markdown(f'<div class="stock-card"><div class="stock-title">{row["Location"]}</div><div class="stock-detail">"{row["Comment"]}"</div></div>',unsafe_allow_html=True)

# ── TAB 7: SKUs ───────────────────────────────────────────────
with tabs[6]:
    s1=k['s1']; s2=k['s2']; s3=k['s3']; tot=s1+s2+s3
    if tot>0:
        sec("SKU / Flavor Performance", "purple", "🥤")
        kpi_row([
            (f"{s1:.0f}", "PINEAPPLE GUAVA", "gold", f"{s1/tot*100:.1f}% share"),
            (f"{s2:.0f}", "LEMON LIME", "teal", f"{s2/tot*100:.1f}% share"),
            (f"{s3:.0f}", "ORANGE MANGO", "purple", f"{s3/tot*100:.1f}% share"),
        ])
        fig, axs = setup_fig(1,2,12,4.5)
        sc_='#A8C4E0','#1F2260','#8B7362'
        axs[0].set_facecolor('#111111')
        axs[0].pie([s1,s2,s3],labels=['Pineapple Guava','Lemon Lime','Orange Mango'],
               colors=sc_,autopct='%1.1f%%',startangle=90,
               textprops={'color':WHITE,'fontsize':10})
        axs[0].set_title('Unit Share by Flavor',pad=10)
        bars=axs[1].bar(['Pineapple Guava','Lemon Lime','Orange Mango'],[s1,s2,s3],
                        color=sc_,edgecolor='none',width=0.6)
        axs[1].set_title('Total Units Sold by SKU',pad=10)
        blbl(axs[1],bars,'{:.0f}',fs=11,color=WHITE)
        plt.tight_layout(); st.pyplot(fig)
        charts_excel['SKU'] = chart_buf(fig)
    else:
        st.info("SKU columns not mapped. Edit Column Mapping to add them.")

# ── TAB 8: CAMPAIGN COMPARE ───────────────────────────────────
with tabs[7]:
    if r['_campaign'].nunique()>1:
        sec("Campaign vs Campaign", "gold", "📊")
        rows=[]
        for c in r['_campaign'].unique():
            rc=r[r['_campaign']==c]; kc=compute_kpis(rc)
            rows.append({'Campaign':c,'Events':kc['n'],'Units':f"{kc['units']:.0f}",
                'Samples':f"{kc['samp']:.0f}",'Revenue':f"${kc['rev']:,.0f}",
                'Conv %':f"{kc['conv']}%",'Avg Units/Demo':kc['avg_u'],
                'Avg Sampled':kc['avg_s'],'Units/Hour':kc['uph']})
        comp=pd.DataFrame(rows)
        st.dataframe(comp,use_container_width=True,hide_index=True)

        fig, axs = setup_fig(1,3,14,4)
        for ax,(col,title,clr) in zip(axs,[('Avg Units/Demo','Avg Units/Demo',GOLD),
            ('Conv %','Conversion %',TEAL),('Units/Hour','Units/Hour',PURPLE)]):
            vals=[float(str(v).replace('%','')) for v in comp[col]]
            bars=ax.bar(comp['Campaign'],vals,color=clr,edgecolor='none',width=0.6)
            ax.set_title(title,pad=10)
            blbl(ax,bars,'{:.2f}',fs=10,color=WHITE)
            plt.setp(ax.xaxis.get_majorticklabels(),rotation=20,ha='right',fontsize=8)
        plt.tight_layout(); st.pyplot(fig)
    else:
        st.info("Upload or select more than one campaign to see the comparison.")

# ── TAB 9: DOWNLOAD ───────────────────────────────────────────
with tabs[8]:
    sec("Download Full Report", "gold", "📥")
    st.markdown(f'<div style="color:{LGRAY};font-size:0.82rem;margin-bottom:16px;">All data tables and charts embedded in the Excel file.</div>',unsafe_allow_html=True)

    def build_excel_report(r, m, charts):
        buf=io.BytesIO()
        ORG_F=PatternFill('solid',fgColor='F0B43C')
        DRK_F=PatternFill('solid',fgColor='141414')
        ALT1=PatternFill('solid',fgColor='1C1C1C')
        ALT2=PatternFill('solid',fgColor='242424')
        thin=Border(left=Side(style='thin',color='2A2A2A'),right=Side(style='thin',color='2A2A2A'),
                    top=Side(style='thin',color='2A2A2A'),bottom=Side(style='thin',color='2A2A2A'))

        k2=compute_kpis(r)
        sheets={}

        sheets['KPI Summary']=pd.DataFrame([
            ['Total Events',k2['n'],'COUNT of all rows'],
            ['Total Samples',f"{k2['samp']:,.0f}",'SUM Sampled'],
            ['Total Units Sold',f"{k2['units']:,.0f}",'SUM Units'],
            ['Total Revenue',f"${k2['rev']:,.2f}",'SUM Revenue'],
            ['Conversion Rate',f"{k2['conv']}%",'Units/Sampled x100'],
            ['Avg Units/Demo',k2['avg_u'],'Units/Events'],
            ['Avg Sampled/Demo',k2['avg_s'],'Sampled/Events'],
            ['Units/BA Hour',k2['uph'],'Units/Hours'],
            ['Coupon Lift',f"+{round((k2['wc']/k2['nc']-1)*100,1)}%" if k2['nc']>0 else 'N/A','(With/Without-1)x100'],
            ['Stock Issue Events',f"{k2['stock_e']} ({stock_pct}%)","Events with inventory concern"],
            ['Sell-Outs',k2['soldout'],'Events confirmed sold out'],
        ],columns=['Metric','Value','How Calculated'])

        raw_cols=[m[v] for v in ['date','ba_name','venue','city','state','units','revenue','sampled','hours','coupons'] if m.get(v)]
        cl=r[raw_cols].copy(); cl['Conv %']=r['_conv'].values
        sheets['All Events']=cl

        sg2=r.groupby('_state').agg(n=('_units','size'),u=('_units','sum'),
            s=('_sampled','sum'),rev=('_revenue','sum')).reset_index()
        sg2['avg']=(sg2.u/sg2.n).round(2); sg2['conv']=(sg2.u/sg2.s*100).round(2)
        sg2.columns=['State','Events','Units','Sampled','Revenue','Avg/Demo','Conv %']
        sheets['By State']=sg2.sort_values('Units',ascending=False)

        ba2=r.groupby('_ba').agg(n=('_units','size'),u=('_units','sum'),s=('_sampled','sum')).reset_index()
        ba2['avg']=(ba2.u/ba2.n).round(2); ba2['conv']=(ba2.u/ba2.s*100).round(2)
        ba2.columns=['BA Name','Events','Units','Sampled','Avg/Demo','Conv %']
        sheets['By BA']=ba2.sort_values('Avg/Demo',ascending=False)

        dow2=r.groupby('_dow').agg(n=('_units','size'),u=('_units','sum'),s=('_sampled','sum')).reset_index()
        dow2['avg']=(dow2.u/dow2.n).round(2); dow2['conv']=(dow2.u/dow2.s*100).round(2)
        dow2['s2']=dow2['_dow'].map({d:i for i,d in enumerate(['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'])})
        dow2=dow2.sort_values('s2').drop('s2',axis=1)
        dow2.columns=['Day','Events','Units','Sampled','Avg/Demo','Conv %']
        sheets['By Day']=dow2

        st2=r.groupby(['_venue','_city','_state']).agg(n=('_units','size'),u=('_units','sum'),s=('_sampled','sum')).reset_index()
        st2['avg']=(st2.u/st2.n).round(2); st2['conv']=(st2.u/st2.s*100).round(2)
        st2.columns=['Store','City','State','Events','Units','Sampled','Avg/Demo','Conv %']
        sheets['Top 20 Stores']=st2.sort_values('Units',ascending=False).head(20)
        sheets['Low 20 Stores']=st2.sort_values('Units').head(20)

        stock_df=r[r['_stock_issue']][['_ba','_city','_state','_units','_sampled','_sold_out','_low_stock']].copy()
        stock_df.columns=['BA','City','State','Units','Sampled','Sold Out','Low Stock']
        sheets['Stock Issues']=stock_df.sort_values('Units',ascending=False)

        r3=r.copy(); r3['tier']=pd.cut(r3['_sampled'],bins=[0,50,65,80,300],labels=['≤50','51-65','66-80','80+'])
        sg3=r3.groupby('tier',observed=True).agg(n=('_units','size'),u=('_units','sum'),s=('_sampled','sum')).reset_index()
        sg3['avg']=(sg3.u/sg3.n).round(2); sg3.columns=['Tier','Events','Units','Sampled','Avg/Demo']
        sheets['Sampling Tiers']=sg3

        s1=k2['s1']; s2=k2['s2']; s3=k2['s3']; tot=s1+s2+s3
        if tot>0:
            sheets['SKU']=pd.DataFrame({'SKU':['Pineapple Guava','Lemon Lime','Orange Mango'],
                'Units':[s1,s2,s3],'Share %':[round(s1/tot*100,1),round(s2/tot*100,1),round(s3/tot*100,1)]})

        with pd.ExcelWriter(buf,engine='openpyxl') as writer:
            for sn,df in sheets.items(): df.to_excel(writer,sheet_name=sn,index=False)

        buf.seek(0); wb=load_workbook(buf)
        from openpyxl.drawing.image import Image as XLImg
        for sn,ibuf in charts.items():
            ws=wb[sn] if sn in wb.sheetnames else wb.create_sheet(sn)
            ibuf.seek(0); img=XLImg(ibuf); img.width=900; img.height=350
            ws.add_image(img,f'A{ws.max_row+3}')

        gold_f=PatternFill('solid',fgColor='F0B43C')
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill=gold_f; cell.font=Font(bold=True,color='000000',size=10)
                cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
                cell.border=thin
            ws.row_dimensions[1].height=26
            for i,row in enumerate(ws.iter_rows(min_row=2),1):
                bg=ALT1 if i%2 else ALT2
                for cell in row:
                    cell.fill=bg; cell.border=thin
                    cell.alignment=Alignment(vertical='center',wrap_text=True)
                    cell.font=Font(color='EEEEEE',size=9)
            for j in range(1,ws.max_column+1):
                ws.column_dimensions[get_column_letter(j)].width=20
            ws.freeze_panes='A2'

        final=io.BytesIO(); wb.save(final); final.seek(0); return final

    with st.spinner("Building report..."):
        excel=build_excel_report(r,m,charts_excel)

    st.download_button("⬇️  Download Full Excel Report", data=excel,
        file_name="Mariano_Events_Campaign_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.markdown("<br>", unsafe_allow_html=True)
    items=["✅ KPI Summary","✅ All Events","✅ By State + chart","✅ By BA Ambassador + chart",
           "✅ By Day of Week + chart","✅ Top 20 Stores","✅ Low 20 Stores","✅ Stock Issues",
           "✅ Sampling Tiers + chart","✅ SKU Breakdown + chart","✅ Coupon Analysis + chart"]
    for item in items:
        st.markdown(f'<div style="color:{LGRAY};font-size:0.8rem;padding:2px 0;">{item}</div>',unsafe_allow_html=True)
